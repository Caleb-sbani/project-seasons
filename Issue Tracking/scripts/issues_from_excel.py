import openpyxl
import glob

directory = input("input directory: ")
search = directory+'/Issues.xlsx'
print("searching for all files following format: " + search)
print("opening files")
files = glob.glob(search)
if files.__len__() != 1:
    print("too many or too lil files")
    print(files)
    exit(1)

table1 = openpyxl.load_workbook(files[0])
table2 = table1.active

listTable = []

class Issue():
    def __init__(self, issueName, issueID, timeline, Tags, Description, takers):
        self.issueName = issueName
        self.issueID = issueID
        self.timeline = timeline
        self.Tags = Tags
        self.Description = Description
        self.takers = takers

for row in table2.iter_rows(2, table2.max_row):
        listTable.append(Issue(*[row[col].value for col in range(0,6)]))
writeLines = []
for issue in listTable:
    writeLines.append(f"- id: {issue.issueID}")
    writeLines.append(f"- title: {issue.issueName}")
    writeLines.append(f"- body: \"{issue.Description}\"")
    tags = issue.Tags.split(", ")
    writeLines.append(f"- labels: ")
    for tag in tags:
        writeLines.append(f"  - {tag}")
    writeLines.append(f"  - {issue.timeline}")
with open(f"{directory}/header.yml") as f:
    header = f.readlines()
    f.close()

with open(f"{directory}/issues.yml", "w") as f:
    f.writelines(header)
    f.writelines([line+"\n" for line in writeLines])
    f.close()

